import csv
from io import BytesIO, StringIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from employment.models import Employee
from users.models import FreyjaUser


class OrganizationChartExportApiTests(TestCase):
    def setUp(self) -> None:
        self.viewer = self._create_employee("viewer@example.com", "Chart", "Viewer")
        self.root = self._create_employee("root@example.com", "Root", "Manager")
        self.child = self._create_employee(
            "child@example.com", "Direct", "Report", manager=self.root
        )
        self.grandchild = self._create_employee(
            "grandchild@example.com", "Nested", "Report", manager=self.child
        )
        self.other_root = self._create_employee("other@example.com", "Other", "Root")
        self.client.force_login(self.viewer.user)

    def test_exports_the_whole_active_organization(self) -> None:
        response = self.client.get(reverse("organization_chart_export"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        rows = self._rows(response.content)
        self.assertEqual(rows[0], ["Level", "Employee", "Job title", "Email", "Manager"])
        self.assertEqual(
            {row[3] for row in rows[1:]},
            {employee.user.email for employee in (self.viewer, self.root, self.child, self.grandchild, self.other_root)},
        )

    def test_exports_only_the_selected_subtree(self) -> None:
        response = self.client.get(
            reverse("organization_chart_subtree_export", args=(self.root.id,))
        )

        self.assertEqual(response.status_code, 200)
        rows = self._rows(response.content)
        self.assertEqual([row[3] for row in rows[1:]], [
            self.root.user.email,
            self.child.user.email,
            self.grandchild.user.email,
        ])
        self.assertEqual([row[0] for row in rows[1:]], ["0", "1", "2"])

    def test_exports_the_selected_subtree_as_xlsx(self) -> None:
        response = self.client.get(
            reverse("organization_chart_subtree_export", args=(self.root.id,)),
            {"file_format": "xlsx"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook["Organization chart"].iter_rows(values_only=True))
        self.assertEqual([row[3] for row in rows[1:]], [
            self.root.user.email,
            self.child.user.email,
            self.grandchild.user.email,
        ])

    def test_rejects_an_unknown_export_format(self) -> None:
        response = self.client.get(
            reverse("organization_chart_export"), {"file_format": "pdf"}
        )

        self.assertEqual(response.status_code, 400)

    def _create_employee(
        self,
        email: str,
        first_name: str,
        last_name: str,
        manager: Employee | None = None,
    ) -> Employee:
        user = FreyjaUser.objects.create_user(
            username=email,
            email=email,
            password="test-password",
            first_name=first_name,
            last_name=last_name,
        )
        return Employee.objects.create(user=user, manager=manager, job_title="Tester")

    def _rows(self, content: bytes) -> list[list[str]]:
        return list(csv.reader(StringIO(content.decode("utf-8-sig"))))
